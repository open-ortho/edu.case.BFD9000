from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from archive.constants import SYSTEM_IDENTIFIER_BOLTON_SUBJECT, SYSTEM_IDENTIFIER_BRUSH
from archive.models import Collection, Coding, Identifier, Subject


@dataclass
class ImportStats:
    subjects_created: int = 0
    subjects_updated: int = 0
    identifiers_created: int = 0
    identifiers_attached: int = 0
    skeletal_missing: int = 0
    rows_skipped: int = 0


class Command(BaseCommand):
    help = "Import Bolton subjects and identifiers from BoltonSubjects2.xlsx"

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--file",
            default="BoltonSubjects2.xlsx",
            help="Path to BoltonSubjects2.xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate without writing to the database",
        )

    def handle(self, *args, **options) -> None:
        try:
            import openpyxl
        except ImportError as exc:
            raise CommandError("openpyxl is required to read .xlsx files") from exc

        file_path = Path(options["file"]).expanduser().resolve()
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        class_codes = self._build_class_codes()
        coding_cache = self._load_coding_cache()

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if "Sheet1" not in wb.sheetnames:
                raise CommandError("Expected Sheet1 in workbook")

            ws = wb["Sheet1"]
            rows = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows, None)
            if not header:
                raise CommandError("Missing header row")

            index = self._build_header_index(header)
            stats = ImportStats()

            if options["dry_run"]:
                self.stdout.write("Dry run enabled: no database writes will occur.")

            with transaction.atomic():
                for row in rows:
                    if not row or all(cell is None for cell in row):
                        continue
                    try:
                        self._import_row(row, index, class_codes, coding_cache, stats)
                    except CommandError as exc:
                        stats.rows_skipped += 1
                        self.stderr.write(str(exc))

                if options["dry_run"]:
                    transaction.set_rollback(True)
                    self.stdout.write("Dry run complete. Rolling back transaction.")
                    self._print_summary(stats)
                    return

            self._print_summary(stats)
        finally:
            wb.close()

    def _build_header_index(self, header: Tuple[object, ...]) -> Dict[str, int]:
        normalized = [str(value).strip().lower() if value is not None else "" for value in header]
        required = [
            "collectionid",
            "subjectid",
            "sex",
            "ethnicitycode",
            "birthdate",
            "angleclass",
            "brushid",
        ]
        missing = [name for name in required if name not in normalized]
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(missing)}")
        return {name: normalized.index(name) for name in required}

    def _build_class_codes(self) -> Dict[str, Optional[str]]:
        return {
            "Class I": "248292005",
            "Class II": "248293000",
            "Class III": "248294006",
            "NULL": None,
        }

    def _load_coding_cache(self) -> Dict[Tuple[str, str], Coding]:
        skeletal_system = "http://snomed.info/sct"
        race_system = "urn:oid:2.16.840.1.113883.6.238"
        skeletal_codes = ["248292005", "248293000", "248294006"]
        race_codes = ["2106-3", "2054-5"]

        codings = Coding.objects.filter(
            system__in=[skeletal_system, race_system],
            code__in=skeletal_codes + race_codes,
        )
        cache = {(coding.system, coding.code): coding for coding in codings}

        missing = []
        for code in skeletal_codes:
            if (skeletal_system, code) not in cache:
                missing.append(f"{skeletal_system}#{code}")
        for code in race_codes:
            if (race_system, code) not in cache:
                missing.append(f"{race_system}#{code}")

        if missing:
            raise CommandError(
                "Missing required Coding entries. Run migrations to seed codes. "
                f"Missing: {', '.join(missing)}"
            )

        return cache

    def _import_row(
        self,
        row: Tuple[object, ...],
        index: Dict[str, int],
        class_codes: Dict[str, Optional[str]],
        coding_cache: Dict[Tuple[str, str], Coding],
        stats: ImportStats,
    ) -> None:
        collection_id = self._cell_str(row[index["collectionid"]])
        subject_id = self._cell_str(row[index["subjectid"]])
        sex_value = self._cell_str(row[index["sex"]])
        ethnicity_value = row[index["ethnicitycode"]]
        birth_date = row[index["birthdate"]]
        angle_class = self._cell_str(row[index["angleclass"]])
        brush_id = self._cell_str(row[index["brushid"]])

        if not collection_id or not subject_id or not birth_date or not sex_value:
            raise CommandError(f"Skipping row with missing required values: {subject_id}")

        collection, _ = Collection.objects.get_or_create(
            short_name=collection_id,
            defaults={"full_name": collection_id},
        )

        subject = self._get_subject_by_identifier(subject_id)
        is_new_subject = subject is None
        if subject is None:
            subject = Subject.objects.create(
                gender=self._map_gender(sex_value),
                birth_date=self._normalize_date(birth_date),
                collection=collection,
            )
            stats.subjects_created += 1
        else:
            updated = False
            gender = self._map_gender(sex_value)
            birth_date_value = self._normalize_date(birth_date)
            if subject.gender != gender:
                subject.gender = gender
                updated = True
            if subject.birth_date != birth_date_value:
                subject.birth_date = birth_date_value
                updated = True
            if subject.collection_id is None:
                subject.collection = collection
                updated = True
            if updated:
                subject.save()
                stats.subjects_updated += 1

        ethnicity = self._resolve_race(ethnicity_value, coding_cache)
        if ethnicity and subject.ethnicity_id != ethnicity.id:
            subject.ethnicity = ethnicity
            subject.save()

        skeletal = self._resolve_skeletal_pattern(angle_class, class_codes, coding_cache)
        if skeletal:
            if subject.skeletal_pattern_id != skeletal.id:
                subject.skeletal_pattern = skeletal
                subject.save()
        else:
            stats.skeletal_missing += 1

        self._attach_identifier(
            subject,
            SYSTEM_IDENTIFIER_BOLTON_SUBJECT,
            subject_id,
            use="official",
            stats=stats,
        )
        if brush_id:
            self._attach_identifier(
                subject,
                SYSTEM_IDENTIFIER_BRUSH,
                brush_id,
                use="secondary",
                stats=stats,
            )

        if not is_new_subject and subject.pk:
            subject.save()

    def _get_subject_by_identifier(self, value: str) -> Optional[Subject]:
        return (
            Subject.objects.filter(
                identifiers__system=SYSTEM_IDENTIFIER_BOLTON_SUBJECT,
                identifiers__value=value,
            )
            .distinct()
            .first()
        )

    def _attach_identifier(
        self,
        subject: Subject,
        system: str,
        value: str,
        use: str,
        stats: ImportStats,
    ) -> None:
        identifier, created = Identifier.objects.get_or_create(
            system=system,
            value=value,
            defaults={"use": use},
        )
        if created:
            stats.identifiers_created += 1
        if not subject.identifiers.filter(pk=identifier.pk).exists():
            subject.identifiers.add(identifier)
            stats.identifiers_attached += 1

    def _resolve_skeletal_pattern(
        self,
        angle_class: str,
        class_codes: Dict[str, Optional[str]],
        coding_cache: Dict[Tuple[str, str], Coding],
    ) -> Optional[Coding]:
        if not angle_class:
            return None
        normalized = angle_class.strip()
        code = class_codes.get(normalized)
        if not code:
            return None
        system = "http://snomed.info/sct"
        return coding_cache.get((system, code))

    def _map_gender(self, value: str) -> str:
        gender_map = {label[0].upper(): key for key, label in Subject.GENDER_CHOICES}
        normalized = value.strip().upper()
        return gender_map.get(normalized, "unknown")

    def _resolve_race(self, value, coding_cache: Dict[Tuple[str, str], Coding]) -> Optional[Coding]:
        race_map = {
            "1": "2106-3",
            "0": "2054-5",
        }
        normalized = str(value).strip()
        code = race_map.get(normalized)
        if not code:
            return None
        system = "urn:oid:2.16.840.1.113883.6.238"
        return coding_cache.get((system, code))

    def _normalize_date(self, value) -> date:
        if isinstance(value, datetime):
            return value.date()
        if hasattr(value, "date"):
            return value.date()
        try:
            return datetime.strptime(str(value), "%Y-%m-%d").date()
        except ValueError as exc:
            raise CommandError(f"Invalid date format: {value}") from exc

    def _cell_str(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _print_summary(self, stats: ImportStats) -> None:
        self.stdout.write("Import complete")
        self.stdout.write(f"Subjects created: {stats.subjects_created}")
        self.stdout.write(f"Subjects updated: {stats.subjects_updated}")
        self.stdout.write(f"Identifiers created: {stats.identifiers_created}")
        self.stdout.write(f"Identifiers attached: {stats.identifiers_attached}")
        self.stdout.write(f"Rows skipped: {stats.rows_skipped}")
        self.stdout.write(f"Skeletal patterns missing: {stats.skeletal_missing}")
