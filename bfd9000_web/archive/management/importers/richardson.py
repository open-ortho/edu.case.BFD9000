"""Importer for the Richardson Collection xlsx spreadsheet."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Optional, Tuple

from django.core.management.base import CommandError
from django.db import transaction

from archive.constants import (
    SYSTEM_IDENTIFIER_RICHARDSON_OLD,
    SYSTEM_IDENTIFIER_RICHARDSON_SUBJECT,
    SYSTEM_RECORD_TYPE,
)
from archive.management.importers.base import BaseImporter, ImportStats
from archive.models import Coding, Encounter, PhysicalRecord, Subject

COLLECTION_SHORT_NAME = 'Richardson'
COLLECTION_FULL_NAME = 'Richardson Collection'

# Map (modality.lower(), projection.lower()) -> record_type code.
# Modality and projection values are from the Terms sheet and observed data.
RICHARDSON_RECORD_TYPE_MAP: Dict[Tuple[str, str], str] = {
    # Study models — projection is typically "None" (the string) or blank
    ('study models', 'none'): 'SM',
    ('study models', ''): 'SM',
    # Radiographs by projection
    ('radiographs', 'lateral'): 'L',
    ('radiographs', 'frontal/pa'): 'F',
    ('radiographs', 'oblique'): 'OB',
    ('radiographs', 'hand/wrist'): 'H',
    ('radiographs', 'occlusal'): 'OC',
    ('radiographs', 'occlusal '): 'OC',   # trailing-space variant seen in data
    ('radiographs', 'none'): 'UK',        # projection unknown
    ('radiographs', ''): 'UK',
    # Photographs
    ('picture', 'lateral'): 'PH',
    ('picture', 'none'): 'PH',
    ('picture', ''): 'PH',
    ('picture', 'occlusal'): 'PH',
    # Cephalometric tracings
    ('tracings', 'lateral'): 'RT',
    ('tracings', 'none'): 'RT',
    ('tracings', ''): 'RT',
}

REQUIRED_RECORD_TYPE_CODES = ['SM', 'L', 'F', 'OB', 'H', 'OC', 'UK', 'PH', 'RT']


@dataclass
class RichardsonStats(ImportStats):
    """Import counters specific to the Richardson dataset."""
    encounters_created: int = 0
    encounters_skipped: int = 0
    physical_records_created: int = 0
    physical_records_skipped: int = 0


class RichardsonImporter(BaseImporter):
    """Import Richardson subjects, encounters, and physical records from an xlsx workbook."""

    def __init__(
        self,
        *,
        dry_run: bool,
        include_names: bool,
        stdout,
        stderr,
    ) -> None:
        super().__init__(
            dry_run=dry_run,
            include_names=include_names,
            stdout=stdout,
            stderr=stderr,
        )

    def run(self, file_path: Path) -> None:
        """Execute the Richardson import from the provided workbook path."""
        try:
            import openpyxl  # pylint: disable=import-outside-toplevel
        except ImportError as exc:
            raise CommandError("openpyxl is required to read .xlsx files") from exc

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        record_type_cache = self._load_record_type_cache()
        class_codes = self._build_class_codes()
        skeletal_cache = self._load_skeletal_coding_cache()
        collection = self._get_or_create_collection(COLLECTION_SHORT_NAME, COLLECTION_FULL_NAME)
        procedure_code = self._get_or_create_procedure()

        stats = RichardsonStats()

        if self.dry_run:
            self.stdout.write("Dry run enabled: no database writes will occur.")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if 'Subject Info' not in wb.sheetnames:
                raise CommandError("Expected 'Subject Info' sheet in workbook")
            if 'Main' not in wb.sheetnames:
                raise CommandError("Expected 'Main' sheet in workbook")

            with transaction.atomic():
                # --- Subject Info sheet ---
                # Row layout: rows 0-3 are title/header rows; data starts at index 4
                ws_subjects = wb['Subject Info']
                subject_rows = list(ws_subjects.iter_rows(values_only=True))
                for row in subject_rows[4:]:
                    if row is None or all(cell is None for cell in row):
                        continue
                    try:
                        self._import_subject_row(row, collection, class_codes, skeletal_cache, stats)
                    except CommandError as exc:
                        stats.rows_skipped += 1
                        self.stderr.write(str(exc))

                # Build an in-memory map of r_number -> Subject for fast Main sheet lookups
                subject_map = self._build_subject_map()

                # --- Main sheet ---
                # Rows 0-1 are partial header rows; data starts at index 2
                ws_main = wb['Main']
                main_rows = list(ws_main.iter_rows(values_only=True))
                for row in main_rows[2:]:
                    if row is None or all(cell is None for cell in row):
                        continue
                    try:
                        self._import_main_row(row, subject_map, record_type_cache, procedure_code, stats)
                    except CommandError as exc:
                        stats.rows_skipped += 1
                        self.stderr.write(str(exc))

                if self.dry_run:
                    transaction.set_rollback(True)
                    self.stdout.write("Dry run complete. Rolling back transaction.")
                    self._print_summary(stats)
                    return

        finally:
            wb.close()

        self._print_summary(stats)

    # ------------------------------------------------------------------
    # Record-type cache
    # ------------------------------------------------------------------

    def _load_record_type_cache(self) -> Dict[str, Coding]:
        """Load required record_type Coding objects; raise if any are missing."""
        codings = Coding.objects.filter(
            system=SYSTEM_RECORD_TYPE,
            code__in=REQUIRED_RECORD_TYPE_CODES,
        )
        cache: Dict[str, Coding] = {c.code: c for c in codings}
        missing = [code for code in REQUIRED_RECORD_TYPE_CODES if code not in cache]
        if missing:
            raise CommandError(
                "Missing required record type Coding entries. "
                "Run 'python manage.py import_valuesets' first. "
                f"Missing codes: {', '.join(missing)}"
            )
        return cache

    # ------------------------------------------------------------------
    # Subject Info import
    # ------------------------------------------------------------------

    def _import_subject_row(
        self,
        row: Tuple,
        collection,
        class_codes: Dict[str, Optional[str]],
        skeletal_cache: Dict[Tuple[str, str], Coding],
        stats: RichardsonStats,
    ) -> None:
        r_number = self._cell_str(row[0])
        if not r_number:
            return  # blank row

        birth_date_raw = row[2]
        old_id = self._cell_str(row[3])
        misc = self._cell_str(row[4])
        gender_raw = self._cell_str(row[5])
        molar_class_raw = self._cell_str(row[6])  # "I", "II", or "III"

        if not birth_date_raw:
            self.stderr.write(f"Skipping subject {r_number}: missing birth date")
            stats.rows_skipped += 1
            return

        birth_date = self._normalize_date(birth_date_raw)
        gender = self._map_gender(gender_raw) if gender_raw else 'unknown'

        # Prepend "Class " so the label matches BaseImporter._build_class_codes keys
        molar_label = f"Class {molar_class_raw.strip()}" if molar_class_raw.strip() else ''
        skeletal = self._resolve_skeletal_pattern(molar_label, class_codes, skeletal_cache)

        # Parse name from "FAMILY, GIVEN" format in col 1
        humanname_family: Optional[str] = None
        humanname_given: Optional[str] = None
        if self.include_names:
            name_raw = self._cell_str(row[1])
            if name_raw:
                parts = name_raw.split(',', 1)
                humanname_family = parts[0].strip() or None
                humanname_given = parts[1].strip() if len(parts) > 1 else None

        subject = self._get_subject_by_r_number(r_number)
        if subject is None:
            subject = Subject.objects.create(
                gender=gender,
                birth_date=birth_date,
                collection=collection,
                skeletal_pattern=skeletal,
                notes=misc,
                humanname_family=humanname_family,
                humanname_given=humanname_given,
            )
            stats.subjects_created += 1
        else:
            updated = False
            if gender and subject.gender != gender:
                subject.gender = gender
                updated = True
            if subject.birth_date != birth_date:
                subject.birth_date = birth_date
                updated = True
            if subject.collection_id is None:  # type: ignore[attr-defined]
                subject.collection = collection
                updated = True
            if skeletal and subject.skeletal_pattern_id != skeletal.id:  # type: ignore[attr-defined]
                subject.skeletal_pattern = skeletal
                updated = True
            if misc and subject.notes != misc:
                subject.notes = misc
                updated = True
            if self.include_names:
                if humanname_family and subject.humanname_family != humanname_family:
                    subject.humanname_family = humanname_family
                    updated = True
                if humanname_given and subject.humanname_given != humanname_given:
                    subject.humanname_given = humanname_given
                    updated = True
            if updated:
                subject.save()
                stats.subjects_updated += 1

        self._attach_identifier(
            subject,
            SYSTEM_IDENTIFIER_RICHARDSON_SUBJECT,
            r_number,
            use='official',
            stats=stats,
        )
        if old_id:
            self._attach_identifier(
                subject,
                SYSTEM_IDENTIFIER_RICHARDSON_OLD,
                old_id,
                use='secondary',
                stats=stats,
            )

    def _get_subject_by_r_number(self, r_number: str) -> Optional[Subject]:
        return (
            Subject.objects.filter(
                identifiers__system=SYSTEM_IDENTIFIER_RICHARDSON_SUBJECT,
                identifiers__value=r_number,
            )
            .distinct()
            .first()
        )

    def _build_subject_map(self) -> Dict[str, Subject]:
        """Build r_number -> Subject mapping for all Richardson subjects."""
        subjects = (
            Subject.objects.filter(
                identifiers__system=SYSTEM_IDENTIFIER_RICHARDSON_SUBJECT,
            )
            .prefetch_related('identifiers')
            .distinct()
        )
        result: Dict[str, Subject] = {}
        for subject in subjects:
            for identifier in subject.identifiers.all():
                if identifier.system == SYSTEM_IDENTIFIER_RICHARDSON_SUBJECT:
                    result[identifier.value] = subject
        return result

    # ------------------------------------------------------------------
    # Main sheet import
    # ------------------------------------------------------------------

    def _import_main_row(
        self,
        row: Tuple,
        subject_map: Dict[str, Subject],
        record_type_cache: Dict[str, Coding],
        procedure_code: Coding,
        stats: RichardsonStats,
    ) -> None:
        r_number = self._cell_str(row[2])
        if not r_number:
            return

        modality_raw = self._cell_str(row[6])
        projection_raw = self._cell_str(row[7])

        # Skip spurious header-repeat rows embedded in the data
        if modality_raw.lower() == 'modality' or projection_raw.lower() == 'projection':
            return

        acq_date_raw = row[8]
        if not acq_date_raw:
            return

        box = self._cell_str(row[12] if len(row) > 12 else None)
        misc = self._cell_str(row[13] if len(row) > 13 else None)

        subject = subject_map.get(r_number)
        if subject is None:
            self.stderr.write(f"No subject found for R number {r_number}; skipping record row")
            stats.physical_records_skipped += 1
            return

        try:
            acq_date = self._normalize_date(acq_date_raw)
        except CommandError:
            self.stderr.write(f"Invalid acquisition date for {r_number}: {acq_date_raw!r}; skipping")
            stats.physical_records_skipped += 1
            return

        # One Encounter per (subject, date); create if not yet seen
        encounter, enc_created = Encounter.objects.get_or_create(
            subject=subject,
            actual_period_start=acq_date,
            defaults={
                'actual_period_start_raw': self._cell_str(acq_date_raw),
                'actual_period_start_precision': 'day',
                'procedure_code': procedure_code,
            },
        )
        if enc_created:
            stats.encounters_created += 1
        else:
            stats.encounters_skipped += 1

        # Map (modality, projection) to record_type code; default to UK (unknown)
        mod_key = modality_raw.lower()
        proj_key = projection_raw.lower()
        code = RICHARDSON_RECORD_TYPE_MAP.get((mod_key, proj_key), 'UK')
        record_type = record_type_cache[code]

        # Always create — duplicates (e.g. multiple photos same day) are intentional
        # Spreadsheet dates have no time-of-day or timezone; attach UTC midnight so
        # Django's timezone-aware DateTimeField does not emit a RuntimeWarning.
        acq_datetime: Optional[datetime] = None
        if isinstance(acq_date_raw, datetime):
            acq_datetime = acq_date_raw.replace(tzinfo=timezone.utc)
        PhysicalRecord.objects.create(
            encounter=encounter,
            record_type=record_type,
            acquisition_datetime=acq_datetime,
            physical_location_box=box,
            notes=misc,
        )
        stats.physical_records_created += 1

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------

    def _print_summary(self, stats: RichardsonStats) -> None:
        self.stdout.write("Import complete")
        self.stdout.write(f"  Subjects created:          {stats.subjects_created}")
        self.stdout.write(f"  Subjects updated:          {stats.subjects_updated}")
        self.stdout.write(f"  Identifiers created:       {stats.identifiers_created}")
        self.stdout.write(f"  Identifiers attached:      {stats.identifiers_attached}")
        self.stdout.write(f"  Encounters created:        {stats.encounters_created}")
        self.stdout.write(f"  Encounters already exist:  {stats.encounters_skipped}")
        self.stdout.write(f"  Physical records created:  {stats.physical_records_created}")
        self.stdout.write(f"  Physical records skipped:  {stats.physical_records_skipped}")
        self.stdout.write(f"  Rows skipped:              {stats.rows_skipped}")
