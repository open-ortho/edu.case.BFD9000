"""Importer for the Bolton subject spreadsheet."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple

from django.core.management.base import CommandError
from django.db import transaction

from archive.constants import (
    SYSTEM_IDENTIFIER_BOLTON_SUBJECT,
    SYSTEM_IDENTIFIER_BRUSH,
)
from archive.management.importers.base import BaseImporter, ImportStats
from archive.models import Coding, Subject, Encounter


@dataclass
class BoltonStats(ImportStats):
    """Import counters specific to the Bolton dataset."""
    skeletal_missing: int = 0
    encounters_created: int = 0
    encounters_skipped: int = 0


class BoltonImporter(BaseImporter):
    """Import Bolton subjects and related identifiers/codings."""
    def __init__(
        self,
        *,
        dry_run: bool,
        include_names: bool,
        stdout,
        stderr,
        timepoints_file: Optional[str] = None,
        skip_timepoints: bool = False,
    ) -> None:
        super().__init__(
            dry_run=dry_run, include_names=include_names, stdout=stdout, stderr=stderr
        )
        self.timepoints_file: Optional[str] = timepoints_file
        self.skip_timepoints = skip_timepoints

    def run(self, file_path: Path) -> None:
        """Execute the Bolton import from the provided workbook path."""
        try:
            # Import locally to provide a clear error when missing.
            import openpyxl  # pylint: disable=import-outside-toplevel
        except ImportError as exc:
            raise CommandError("openpyxl is required to read .xlsx files") from exc

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        class_codes = self._build_class_codes()
        coding_cache = self._load_coding_cache()

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            if "Sheet1" not in wb.sheetnames:
                raise CommandError("Expected Sheet1 in workbook")

            ws = wb["Sheet1"]
            rows = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows, None)
            if not header:
                raise CommandError("Missing header row")

            index = self._build_header_index(header)
            stats = BoltonStats()

            if self.dry_run:
                self.stdout.write("Dry run enabled: no database writes will occur.")

            with transaction.atomic():
                procedure_code = self._get_or_create_procedure()
                for row in rows:
                    if not row or all(cell is None for cell in row):
                        continue
                    try:
                        self._import_row(row, index, class_codes, coding_cache, stats)
                    except CommandError as exc:
                        stats.rows_skipped += 1
                        self.stderr.write(str(exc))

                # Optionally import timepoints CSV (generates Encounter rows per subject/timepoint)
                if not self.skip_timepoints:
                    if self.timepoints_file:
                        csv_path = Path(self.timepoints_file).expanduser().resolve()
                    else:
                        csv_path = Path(__file__).resolve().parents[3] / "docs" / "collections_data" / "BoltonTimepoints2.csv"
                    try:
                        self._import_timepoints(csv_path, stats, procedure_code)
                    except CommandError as exc:
                        # Non-fatal for import: report and continue
                        stats.rows_skipped += 1
                        self.stderr.write(str(exc))

                if self.dry_run:
                    transaction.set_rollback(True)
                    self.stdout.write("Dry run complete. Rolling back transaction.")
                    self._print_summary(stats)
                    return

            self._print_summary(stats)
        finally:
            wb.close()

    def _build_header_index(self, header: Tuple[object, ...]) -> Dict[str, int]:
        normalized = [str(value).strip().lower() if value is not None else "" for value in header]
        required = [
            "collectionid",
            "subjectid",
            "sex",
            "ethnicitycode",
            "birthdate",
            "angleclass",
            "brushid",
        ]
        missing = [name for name in required if name not in normalized]
        if missing:
            raise CommandError(f"Missing required columns: {', '.join(missing)}")
        return {name: normalized.index(name) for name in required}

    def _load_coding_cache(self) -> Dict[Tuple[str, str], Coding]:
        skeletal_system = "http://snomed.info/sct"
        race_system = "urn:oid:2.16.840.1.113883.6.238"
        skeletal_codes = ["248292005", "248293000", "248294006"]
        race_codes = ["2106-3", "2054-5"]

        codings = Coding.objects.filter(
            system__in=[skeletal_system, race_system],
            code__in=skeletal_codes + race_codes,
        )
        cache = {(coding.system, coding.code): coding for coding in codings}

        missing = []
        for code in skeletal_codes:
            if (skeletal_system, code) not in cache:
                missing.append(f"{skeletal_system}#{code}")
        for code in race_codes:
            if (race_system, code) not in cache:
                missing.append(f"{race_system}#{code}")

        if missing:
            raise CommandError(
                "Missing required Coding entries. Run migrations to seed codes. "
                f"Missing: {', '.join(missing)}"
            )

        return cache

    def _import_row(
        self,
        row: Tuple[object, ...],
        index: Dict[str, int],
        class_codes: Dict[str, Optional[str]],
        coding_cache: Dict[Tuple[str, str], Coding],
        stats: BoltonStats,
    ) -> None:
        collection_id = self._cell_str(row[index["collectionid"]])
        subject_id = self._cell_str(row[index["subjectid"]])
        sex_value = self._cell_str(row[index["sex"]])
        ethnicity_value = row[index["ethnicitycode"]]
        birth_date = row[index["birthdate"]]
        angle_class = self._cell_str(row[index["angleclass"]])
        brush_id = self._cell_str(row[index["brushid"]])

        if not collection_id or not subject_id or not birth_date or not sex_value:
            raise CommandError(f"Skipping row with missing required values: {subject_id}")

        collection = self._get_or_create_collection(collection_id)

        ethnicity = self._resolve_race(ethnicity_value, coding_cache)
        skeletal = self._resolve_skeletal_pattern(angle_class, class_codes, coding_cache)
        if not skeletal:
            stats.skeletal_missing += 1

        subject = self._get_subject_by_identifier(subject_id)
        if subject is None:
            subject = Subject.objects.create(
                gender=self._map_gender(sex_value),
                birth_date=self._normalize_date(birth_date),
                collection=collection,
                ethnicity=ethnicity,
                skeletal_pattern=skeletal,
            )
            stats.subjects_created += 1
        else:
            updated = False
            gender = self._map_gender(sex_value)
            birth_date_value = self._normalize_date(birth_date)
            if subject.gender != gender:
                subject.gender = gender
                updated = True
            if subject.birth_date != birth_date_value:
                subject.birth_date = birth_date_value
                updated = True
            if subject.collection_id is None:
                subject.collection = collection
                updated = True
            if ethnicity and subject.ethnicity_id != ethnicity.id:
                subject.ethnicity = ethnicity
                updated = True
            if skeletal and subject.skeletal_pattern_id != skeletal.id:
                subject.skeletal_pattern = skeletal
                updated = True
            if updated:
                subject.save()
                stats.subjects_updated += 1

        self._attach_identifier(
            subject,
            SYSTEM_IDENTIFIER_BOLTON_SUBJECT,
            subject_id,
            use="official",
            stats=stats,
        )
        if brush_id:
            self._attach_identifier(
                subject,
                SYSTEM_IDENTIFIER_BRUSH,
                brush_id,
                use="secondary",
                stats=stats,
            )

    def _get_subject_by_identifier(self, value: str) -> Optional[Subject]:
        return (
            Subject.objects.filter(
                identifiers__system=SYSTEM_IDENTIFIER_BOLTON_SUBJECT,
                identifiers__value=value,
            )
            .distinct()
            .first()
        )

    def _import_timepoints(self, csv_path: Path, stats: BoltonStats, procedure_code: Coding) -> None:
        """Read the BoltonTimepoints2.csv and create Encounter objects per timepoint.

        The CSV uses `subjectid` values that correspond to the Bolton identifier
        (system = SYSTEM_IDENTIFIER_BOLTON_SUBJECT). For each row we locate the
        Subject by that identifier and create an Encounter with
        `actual_period_start` populated from `timepointdate`.
        """
        if not csv_path.exists():
            raise CommandError(f"Timepoints CSV not found: {csv_path}")

        with csv_path.open(newline='', encoding='utf-8') as fh:
            reader = csv.DictReader(fh)
            required = ["collectionid", "subjectid", "timepointnum", "timepointdate"]
            # Normalize fieldnames to lowercase so row access is consistent (#7)
            raw_fieldnames = reader.fieldnames or []
            normalized_fieldnames = [h.strip().lower() for h in raw_fieldnames]
            missing = [r for r in required if r not in normalized_fieldnames]
            if missing:
                raise CommandError(f"Bolton timepoints CSV missing columns: {', '.join(missing)}")
            # Build a mapping from normalized key -> original key for DictReader rows
            key_map = {norm: orig for norm, orig in zip(normalized_fieldnames, raw_fieldnames)}

            for raw_row in reader:
                # Normalize row keys to lowercase (#7)
                row = {norm: raw_row[orig] for norm, orig in key_map.items()}
                subject_id = self._cell_str(row.get("subjectid"))
                date_raw = row.get("timepointdate")
                if not subject_id or not date_raw:
                    self.stderr.write(f"Skipping timepoint row with missing subject or date: {row}")
                    stats.encounters_skipped += 1
                    continue

                subject = self._get_subject_by_identifier(subject_id)
                if subject is None:
                    self.stderr.write(f"No subject found for Bolton id {subject_id}; skipping timepoint")
                    stats.encounters_skipped += 1
                    continue

                try:
                    actual_date = self._normalize_date(date_raw)
                except CommandError:
                    self.stderr.write(f"Invalid date for subject {subject_id}: {date_raw}; skipping")
                    stats.encounters_skipped += 1
                    continue

                date_raw_str = self._cell_str(date_raw)  # #8a: use _cell_str instead of str()

                # Avoid creating duplicate encounters for same subject/date/raw (#8b)
                if Encounter.objects.filter(
                    subject=subject,
                    actual_period_start=actual_date,
                    actual_period_start_raw=date_raw_str,
                ).exists():
                    self.stderr.write(f"Encounter already exists for {subject_id} on {actual_date}; skipping")
                    stats.encounters_skipped += 1
                    continue

                # Use objects.create() instead of manual save() (#8c)
                Encounter.objects.create(
                    subject=subject,
                    actual_period_start=actual_date,
                    actual_period_start_raw=date_raw_str,
                    actual_period_start_precision="day",
                    procedure_code=procedure_code,
                )
                stats.encounters_created += 1

    def _resolve_race(self, value, coding_cache: Dict[Tuple[str, str], Coding]) -> Optional[Coding]:
        race_map = {
            "1": "2106-3",
            "0": "2054-5",
        }
        normalized = self._cell_str(value)
        code = race_map.get(normalized)
        if not code:
            return None
        system = "urn:oid:2.16.840.1.113883.6.238"
        return coding_cache.get((system, code))

    def _print_summary(self, stats: BoltonStats) -> None:
        self.stdout.write("Import complete")
        self.stdout.write(f"Subjects created: {stats.subjects_created}")
        self.stdout.write(f"Subjects updated: {stats.subjects_updated}")
        self.stdout.write(f"Identifiers created: {stats.identifiers_created}")
        self.stdout.write(f"Identifiers attached: {stats.identifiers_attached}")
        self.stdout.write(f"Encounters created: {stats.encounters_created}")
        self.stdout.write(f"Encounters skipped: {stats.encounters_skipped}")
        self.stdout.write(f"Rows skipped: {stats.rows_skipped}")
        self.stdout.write(f"Skeletal patterns missing: {stats.skeletal_missing}")
